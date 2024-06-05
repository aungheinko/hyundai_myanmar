import openpyxl

# Load the workbook and select the sheet
file_path = r'D:\MyWorkingEnv\Hyundai\DaiyRo\Service_2R-7663_Warranty_R230125-018.xlsx'
workbook = openpyxl.load_workbook(file_path, data_only=True)
sheet = workbook['Repair_Order_02']

# Initialize variables
data_start = False
header_row = ["RO NUMBER", "Parts No", "Description", "Price ($)", "Price (Ks)", "QTY", "Total Amount($)"]
extracted_data = []

for row in sheet.iter_rows(values_only=True):
    if not data_start:
        # Check if the current row contains "Parts No"
        if "RO-" in row:
            ro_number = row[11]
# Iterate through the rows to find the header and extract data
for row in sheet.iter_rows(values_only=True):
    if not data_start:
        # Check if the current row contains "Parts No"
        if "Parts No" in row:
            data_start = True
    else:
        # If the row contains "Total", stop extracting
        if any(cell is not None and "Total" in str(cell) for cell in row):
            break
        # Append the row to extracted_data if it meets the criteria
        if row[3] is not None or row[4] is not None:
            if row[9] is not None:
                total = row[9] * row[11]
            else:
                total = (row[10]/2100) * row[11]
            extracted_data.append([ro_number,row[2],row[3], row[9], row[10], row[11],total])
            print([ro_number,row[2],row[3], row[9], row[10], row[11],total])
            # print(ro_number,row)
            # extracted_data.append([ro_number,row])
            # extracted_data.append([ro_number,row[3], row[4], row[9], row[10], row[11]])

# Create a new workbook and add a sheet to store the extracted data
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Extracted_Data"

# Write the header row to the new sheet
new_sheet.append(header_row)

# Write the extracted data to the new sheet
for data_row in extracted_data:
    new_sheet.append(data_row)

# Define the new file path
new_file_path = r'D:\MyWorkingEnv\Hyundai\DaiyRo\extracted_data.xlsx'

# Save the new workbook with the extracted data
new_workbook.save(new_file_path)
print(f"Data extraction complete. Saved as '{new_file_path}'")

import os
import openpyxl
import warnings
from tqdm import tqdm

# Suppress specific warning message about conditional formatting
warnings.filterwarnings("ignore", category=UserWarning, message="Conditional Formatting extension is not supported and will be removed")

def extract_data_from_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['Repair_Order_02']

        data_start = False
        extracted_data = []
        ro_number = None

        for row in sheet.iter_rows(values_only=True):
            if not data_start:
                if row and "TEL:  / FAX: " in str(row):
                    ro_number = row[11]
            else:
                if any(cell is not None and "Total" in str(cell) for cell in row):
                    break
                if row[3] is not None or row[4] is not None:
                    if row[9] is not None:
                        total = row[9] * row[11]
                    else:
                        total = (row[10] / 2100) * row[11]
                    extracted_data.append([ro_number, row[2], row[3], row[9], row[10], row[11], total])

            if not data_start:
                if row and "Parts No" in row:
                    data_start = True

        # Filter out rows where ro_number starts with 'Q'
        filtered_data = [data for data in extracted_data if not data[0].startswith('Q')]

        return filtered_data

    except Exception as e:
        print(f"Error processing file {file_path}: {e}")
        return []

def get_files_in_folder(folder_path):
    try:
        files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith(".xlsx")]
        return files
    except FileNotFoundError:
        print("Folder not found.")
        return []

def main():
    folder_path = input("Enter the folder path: ")

    if not os.path.exists(folder_path):
        print("Invalid folder path.")
        return

    file_list = get_files_in_folder(folder_path)
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Extracted_Data"

    header_row = ["RO NUMBER", "Parts No", "Description", "Price ($)", "Price (Ks)", "QTY", "Total Amount($)"]
    new_sheet.append(header_row)

    for file_name in tqdm(file_list, desc="Processing files"):
        extracted_data = extract_data_from_excel(file_name)
        for data_row in extracted_data:
            new_sheet.append(data_row)

    new_file_path = os.path.join(folder_path, 'extracted_data.xlsx')
    new_workbook.save(new_file_path)
    print(f"Data extraction complete. Saved as '{new_file_path}'")

if __name__ == "__main__":
    main()

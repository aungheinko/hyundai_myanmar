import os
from openpyxl import Workbook
from openpyxl import load_workbook
folder_path = rf"D:\Temp\CHECK"

def is_empty_row(row):
    return all(value is None for value in row)

excel_file = rf"D:\Temp\CHECK\QO_9Q-8161_Q240404-008.xlsx"
wb = load_workbook(excel_file)
sheet_name = "Repair_Order_01"
sheet = wb[sheet_name]
# Find the index of the row containing "RO No.:"
ro_row_index = None
for i, row in enumerate(sheet.iter_rows()):
    for cell in row:
        if "RO No.:" in str(cell.value):
            ro_row_index = i
            break
    if ro_row_index is not None:
        break
if ro_row_index is not None:
    # Delete rows above the row containing "RO No.:"
    sheet.delete_rows(1, ro_row_index - 1)
# Define Column Range
column_range = "A:M"
data_list = []
advisor_solve = None
for row in sheet.iter_rows(min_col=1, max_col=13, values_only=True):
    if not is_empty_row(row):
        data_list.append(row)
if data_list[1][10] is not None:
    try:
        car_takein_date = data_list[1][10].strftime("%m/%d/%Y")
    except Exception as e:
        car_takein_date = data_list[1][10]
        pass
try:
    if data_list[15][10] != None:
        advisor_solve = data_list[15][10]
    elif data_list[14][10] != None:
        advisor_solve = data_list[14][10]
    else:
        advisor_solve = None
except Exception as e:
    advisor_solve = None
print(advisor_solve)
print(data_list[11])
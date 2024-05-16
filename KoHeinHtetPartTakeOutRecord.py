import os
from openpyxl import load_workbook, Workbook

def process_excel_files(folder_path, output_file):
    # Create a new workbook to store the combined data
    combined_wb = Workbook()
    combined_ws = combined_wb.active
    
    # Add headers to the combined sheet
    combined_ws.append(['File Name', 'Sequence'] + [f'Column {i+1}' for i in range(10)])  # Assuming 10 columns
    
    # Initialize sequence counter
    sequence = 1
    
    # Iterate over all files in the folder
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):  # Check if file is Excel file
            file_path = os.path.join(folder_path, file_name)
            try:
                # Read the Excel file with data_only=True to get cell values instead of formulas
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb['Sheet1']
                
                # Get the data from the second row
                data = [file_name, sequence] + [cell.value for cell in ws[2]]
                
                # Write the data to the combined sheet
                combined_ws.append(data)
                
                # Increment sequence counter
                sequence += 1
            except Exception as e:
                print(f"Error processing {file_name}: {e}")

    # Save the combined workbook to an output file
    combined_wb.save(output_file)
    print(f"Combined data written to {output_file}")

# Prompt the user to input the folder path
folder_path = input("Enter the folder path containing Excel files: ")

# Check if the provided folder path exists
if not os.path.isdir(folder_path):
    print("Error: Folder path does not exist.")
else:
    output_file = os.path.join(folder_path, 'combined_data.xlsx')
    process_excel_files(folder_path, output_file)

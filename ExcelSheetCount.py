import openpyxl

def count_sheets_and_output_names(input_file, output_file):
    # Open the Excel file
    workbook = openpyxl.load_workbook(input_file)
    
    # Get sheet names
    sheet_names = workbook.sheetnames
    
    # Count the number of sheets
    num_sheets = len(sheet_names)
    
    # Create a new workbook
    new_workbook = openpyxl.Workbook()
    
    # Select the active sheet of the new workbook
    new_sheet = new_workbook.active
    
    # Write sheet names to the new sheet
    for i, sheet_name in enumerate(sheet_names, 1):
        new_sheet.cell(row=i, column=1).value = sheet_name
    
    # Save the new workbook
    new_workbook.save(output_file)
    
    return num_sheets

# Example usage:
input_file = rf"D:\Database\ASD Credit Update_2.xlsx"  # Replace with your input Excel file
output_file = "sheet_names.xlsx"  # Replace with the desired name for the output Excel file

num_sheets = count_sheets_and_output_names(input_file, output_file)
print(f"Number of sheets: {num_sheets}")
print(f"Sheet names written to '{output_file}'.")

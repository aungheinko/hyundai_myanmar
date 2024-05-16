import os
from openpyxl import load_workbook, Workbook

def get_files_in_folder(folder_path):
    try:
        files = os.listdir(folder_path)
        return files
    except FileNotFoundError:
        print("Folder not found.")
        return []

def get_last_value_below_total_amount(filename):
    try:
        wb = load_workbook(filename, data_only=True)
        sheet = wb["Invoice"]
        total_amount_row = None

        # Find the row containing "Total Amount"
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "Total Amount":
                    total_amount_row = cell.row
                    break
            if total_amount_row:
                break

        if total_amount_row:
            next_row = sheet[total_amount_row + 1]
            for cell in reversed(next_row):
                if cell.value is not None:
                    return cell.value
        else:
            print(f"Total Amount not found in the sheet of file: {filename}")
            return None
    except FileNotFoundError:
        print(f"File not found: {filename}")
        return None
    except Exception as e:
        print(f"An error occurred while processing file {filename}: {e}")
        return None

def append_to_excel(file_path, data):
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["RO Number", "Total Amount"])

    for row in data:
        sheet.append(row)
    
    workbook.save(file_path)

# Usage example
folder_path = input("Enter the folder path: ")

if not os.path.exists(folder_path):
    print("Invalid folder path.")
    exit()

file_list = get_files_in_folder(folder_path)
output_data = []

for file_name in file_list:
    excel_file = os.path.join(folder_path, file_name)
    last_underscore_index = excel_file.rfind("_")
    dot_index = excel_file.rfind(".")
    last_value = get_last_value_below_total_amount(excel_file)
    if last_value is not None:
        ro_number = excel_file[last_underscore_index + 1:dot_index]
        total_amount = 0 if isinstance(last_value, str) else last_value
        output_data.append([ro_number, total_amount])

output_file_path = os.path.join(folder_path, "final_invoice_data.xlsx")
append_to_excel(output_file_path, output_data)
print(f"Final data appended to {output_file_path}")

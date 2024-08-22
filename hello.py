import os
import openpyxl
import pandas as pd
import re
import logging

def find_label_value(sheet, label):
    """
    Search for the label in the sheet and return the value from the cell below it.
    """
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell and label in str(cell):
                cell_row = cell.row
                return sheet.cell(row=cell_row + 1, column=cell.column).value
    return None

def extract_data_from_file(file_path, sheet_name):
    try:
        # Load the workbook and select the specific sheet
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name]

        # Find the "Plate No." and "VIN No." labels and extract data from below them
        plate_no = find_label_value(sheet, "Plate No.")
        vin_no = find_label_value(sheet, "VIN No.")

        # Extract data from the specified cells
        merged_cell_LM26 = sheet['L26'].value
        cell_H34 = sheet['H34'].value
        merged_cell_IJ34 = sheet['I34'].value

        return [file_path, plate_no, vin_no, merged_cell_LM26, cell_H34, merged_cell_IJ34]
    
    except Exception as e:
        logging.error(f"Error processing {file_path}: {e}")
        return [file_path, None, None, None, None, None]

def extract_and_append_data(directory_path, output_directory):
    # Get the folder name from the directory path
    folder_name = os.path.basename(os.path.normpath(directory_path))
    output_file = os.path.join(output_directory, f"{folder_name}.xlsx")
    
    # Initialize logging
    logging.basicConfig(filename=os.path.join(output_directory, 'processing.log'), level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')
    
    logging.info(f"Started processing files in directory: {directory_path}")
    
    # Initialize a list to store extracted data
    data_rows = []
    
    # Counter for progress tracking
    file_count = 0
    total_files = len([f for f in os.listdir(directory_path) if f.endswith(".xlsx")])

    # Iterate over all Excel files in the specified directory
    for filename in os.listdir(directory_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory_path, filename)
            
            # Extract data from the current file
            extracted_data = extract_data_from_file(file_path, 'Repair_Order_01')
            data_rows.append(extracted_data)
            
            # Log progress
            file_count += 1
            if file_count % 50 == 0 or file_count == total_files:
                logging.info(f"Processed {file_count} out of {total_files} files")
    
    # Create a DataFrame from the collected data
    columns = ['File Path', 'Plate No.', 'VIN No.', 'L26:M26', 'H34', 'I34:J34']
    df = pd.DataFrame(data_rows, columns=columns)
    
    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Aggregated Data')

    logging.info(f"Data has been extracted and saved to {output_file}")

# Ask the user for the directory path and output directory
directory_path = input("Enter the folder path containing Excel files: ")
output_directory = input("Enter the directory to save the output file: ")

# Run the extraction and appending process
extract_and_append_data(directory_path, output_directory)

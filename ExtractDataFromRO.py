import os
from openpyxl import load_workbook
from openpyxl import Workbook
import warnings
from tqdm import tqdm
import time
from datetime import datetime
import pandas as pd

def ignore_conditional_formatting_warning(message, category, filename, lineno, file=None, line=None):
    pass

# Add the custom filter to ignore the Conditional Formatting warning
warnings.simplefilter("ignore", category=UserWarning)
warnings.formatwarning = ignore_conditional_formatting_warning

warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported")
os.system('cls')

def get_last_value_below_total_amount(filename):
    try:
        wb = load_workbook(filename, data_only=True)  # Use data_only=True to get calculated values instead of formulas
        sheet = wb["Invoice"]
        total_amount_row = None

        # Find the row containing "Total Amount"
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "Total Amount":
                    total_amount_row = cell.row
                    break
            if total_amount_row:
                break

        if total_amount_row:
            total_amount_cell = sheet.cell(row=total_amount_row, column=2)
            next_row = sheet[total_amount_row + 1]
            last_value = None
            for cell in reversed(next_row):
                if cell.value is not None:
                    if cell.data_type == "f":  # Check if the cell contains a formula
                        last_value = cell.value  # If it's a formula, get the formula value
                    else:
                        last_value = cell.value  # Otherwise, get the displayed value
                    break
            return last_value
        else:
            print(f"Total Amount not found in the sheet of file: {filename}")
            return None
    except FileNotFoundError:
        print(f"File not found: {filename}")
        return None
    except Exception as e:
        print(f"An error occurred while processing file {filename}: {e}")
        return None
    
sequence = 0
def calculate_final_total(service_type, excel_file_path):
    workbook = load_workbook(excel_file_path)
    sheet_name = "Repair_Order_02"
    sheet = workbook[sheet_name]
    payment_type = ""

    found_header_parts = False
    for row in sheet.iter_rows(values_only=True):
        if "No" in row and "Parts No" in row and "Description" in row:
            found_header_parts = True
            break
    if not found_header_parts:
        raise ValueError("Header for parts list not found in the sheet.")

    found_header_labour = False
    for row in sheet.iter_rows(values_only=True):
        if "No." in row and "Section" in row and "Description" in row:
            found_header_labour = True
            break
    if not found_header_labour:
        raise ValueError("Header for labour charges list not found in the sheet.")

    header_row_reached_parts = False
    part_takeout_list = []
    for row_tuple in sheet.iter_rows(values_only=True):
        if not header_row_reached_parts:
            if "No" in row_tuple and "Parts No" in row_tuple and "Description" in row_tuple:
                header_row_reached_parts = True
                part_takeout_list.append(row_tuple)
        else:
            if "Total" in row_tuple:
                break
            part_takeout_list.append(row_tuple)

    header_row_reached_labour = False
    labour_charges_list = []
    for row_tuple in sheet.iter_rows(values_only=True):
        if not header_row_reached_labour:
            if "No." in row_tuple and "Section" in row_tuple and "Description" in row_tuple:
                header_row_reached_labour = True
                labour_charges_list.append(row_tuple)
        else:
            if "Total" in row_tuple:
                break
            labour_charges_list.append(row_tuple)

    part_sale_total = 0
    for part in part_takeout_list:
        if part[9] is not None and part[9] != "Price ($)":
            part_sale_total += part[9] * part[11]

    total_labour_charges = 0
    for labour in labour_charges_list:
        if labour[11] is not None and labour[11] != "LTS":
            if service_type == "Warranty":
                total_labour_charges += round(labour[11] * 6, 0)
            else:
                total_labour_charges += labour[11] * 20

    additional_sheet_name = "Invoice"
    additional_sheet = workbook[additional_sheet_name]

    currency_row_index = None
    for row_index, row in enumerate(additional_sheet.iter_rows(values_only=True), start=1):
        if "Currency" in row:
            currency_row_index = row_index
            break

    data_collection = []

    if currency_row_index and currency_row_index < additional_sheet.max_row:
        next_row_values = []
        for cell in additional_sheet[currency_row_index]:
            if cell.data_type == 'f':
                next_row_values.append(cell.value)
            else:
                next_row_values.append(cell.value)
        data_collection.append(next_row_values)

    tax_percentage = 0
    parts_only_discount_amount = 0
    discount_pecent = 0
    parts_and_service_discount = 0
    retail_part = 0
    retail_labour = 0
    part_discount = 0
    labour_discount = 0
    part_tax_percent = 0
    labour_tax_percent = 0
    subtotal = 0

    for cell_value in data_collection[0]:
        if "Tax(5%)" in str(cell_value):
            tax_percentage = float("".join(filter(str.isdigit, str(cell_value))))
        if "parts only" in str(cell_value).lower():
            parts_only_discount_amount = float("".join(filter(str.isdigit, str(cell_value))))
        elif "discount" in str(cell_value).lower():
            discount_pecent = float("".join(filter(str.isdigit, str(cell_value))))
        if "parts & service" in str(cell_value).lower():
            parts_and_service_discount = float("".join(filter(str.isdigit, str(cell_value))))

    if service_type == "Inspection":
        return retail_part,retail_labour,part_discount,labour_discount,part_tax_percent,labour_tax_percent, subtotal
    elif service_type == "Quotation":
        return retail_part,retail_labour,part_discount,labour_discount,part_tax_percent,labour_tax_percent, subtotal
    elif service_type == "Pending":
        return retail_part,retail_labour,part_discount,labour_discount,part_tax_percent,labour_tax_percent, subtotal
    elif service_type == "Warranty":
        return part_sale_total,total_labour_charges,part_discount,labour_discount,part_tax_percent,labour_tax_percent, part_sale_total + total_labour_charges
        print(total_labour_charges + part_sale_total)
    elif service_type == "Auto Parts":
        subtotal = total_labour_charges + part_sale_total
        if tax_percentage != 0:
            tax_total = subtotal * (tax_percentage / 100)
        else:
            tax_total = 0
        total_with_tax = subtotal + tax_total
        if discount_pecent != 0:
            discount_amount = total_with_tax * (discount_pecent / 100)
        else:
            discount_amount = 0
        final_total = total_with_tax - discount_amount
        # print(round(final_total * 2100, -2))
        return part_sale_total,retail_labour,discount_pecent,labour_discount,tax_percentage,tax_percentage,final_total
    elif service_type == "General":
        total_discount = 0
        if parts_and_service_discount != 0:
            total_discount = parts_and_service_discount
        elif parts_only_discount_amount != 0:
            total_discount = parts_only_discount_amount
        elif discount_pecent != 0:
            total_discount = discount_pecent

        subtotal = total_labour_charges + part_sale_total
        if tax_percentage != 0:
            tax_total = subtotal * (tax_percentage / 100)
        else:
            tax_total = 0
        total_with_tax = subtotal + tax_total
        if total_discount != 0:
            discount_amount_with_tax = total_with_tax * (total_discount / 100)
        else:
            discount_amount_with_tax = 0
        final_total = total_with_tax - discount_amount_with_tax
        return part_sale_total, total_labour_charges,total_discount,total_discount, tax_percentage,tax_percentage,final_total
    
def get_files_in_folder(folder_path):
    try:
        # List all files in the folder
        files = os.listdir(folder_path)
        
        # Return the list of file names
        return files
    except FileNotFoundError:
        print("Folder not found.")
        return []

def is_empty_row(row):
    return all(value is None for value in row)

# Prompt the user to input the folder path
folder_path = input("Enter the folder path: ")
# folder_path = rf"D:\RO-Mar_2024\29.03.2024"

# Validate the folder path
if not os.path.exists(folder_path):
    print("Invalid folder path.")
    exit()

file_list = get_files_in_folder(folder_path)
new_file = Workbook()
ws = new_file.active

cash_collection_file = rf"D:\Working Enviroment\Database\Daily Cash Collection.xlsx"
cash_collection_workbook = load_workbook(cash_collection_file)
cash_collection_daily_sheet = cash_collection_workbook['Daily']
cash_collection_data = []

for data_row in cash_collection_daily_sheet.iter_rows(values_only=True,min_col=2,max_col=14):
    if data_row[1] is not None:
        cash_collection_data.append(data_row)

ws.append([
    "RO No", "Take In Date", "Take in Time", "Take out Date", "Take Out Time", "Pending Cars", 
    "Lead Time", "Car Name", "CarType", "Car Maker", "Car No", "Vin No", "Km", "Car Colour", 
    "Car Model yr", "Customer Type", "Customer Name", "Customer Phone No", "ServicePurpose", 
    "Service Type", "Reason", "Repair Type", "Repeat", "Q No", "P No", "V No", "HMSR No", 
    "Retail Part", "Retail Labour", "Part", "Labour", "Part", "Labor", "Part Total", "Labor Total", 
    "Total Amount", "Invoice Amount", "Received Amount", "Remaining Amount", "Payment Type", 
    "Cash/Bank", "Collected?", "Bill Collect Date", "Refund Amount", "Advisor", "Team", "Remark"
])

# progress_bar = tqdm(total=len(file_list))

# Select the active worksheet
for file_name in file_list:
    excel_file = os.path.join(folder_path, file_name)
    if not excel_file.endswith(".xlsx"):
        continue
    else:
        last_value = get_last_value_below_total_amount(excel_file)
        if isinstance(last_value,str):
            last_value = 0
        else:
            try:
                last_value = last_value/2100
            except Exception as e:
                last_value = 0
        wb = load_workbook(excel_file)
        sheet_name = "Repair_Order_01"
        sheet = wb[sheet_name]

        # Find the index of the row containing "RO No.:"
        ro_row_index = None
        for i, row in enumerate(sheet.iter_rows()):
            for cell in row:
                if "RO No.:" in str(cell.value):
                    ro_row_index = i
                    break
            if ro_row_index is not None:
                break

        if ro_row_index is not None:
            # Delete rows above the row containing "RO No.:"
            sheet.delete_rows(1, ro_row_index - 1)

        # Define Column Range
        column_range = "A:M"
        data_list = []

        # Iterate through rows and extract data from columns A to M
        for row in sheet.iter_rows(min_col=1, max_col=13, values_only=True):
            if not is_empty_row(row):
                data_list.append(row)
        if data_list[1][10] is not None:
            try:
                car_takein_date = data_list[1][10].strftime("%m/%d/%Y")
                car_takein_date = car_takein_date.strftime("%-m/%d/%Y")
            except Exception as e:
                car_takein_date = data_list[1][10]
                pass
        else:
            print("Error: car take-in date is None")
            car_takein_date = ""
        car_takeout_date = ""
        car_takein_time = data_list[12][11]
        car_takeout_time = data_list[12][12]
        print(car_takein_time, car_takeout_time)
        time.sleep(3)
        pending = ""
        lead_day = ""
        if car_takeout_time is None:
            car_takeout_date = ""
            pending = "Pending"
            car_takeout_time = ""
        else:
            car_takeout_date = car_takein_date
        if pending == "Pending":
            lead_day = ""
        else:
            lead_day = 1
        #Ro Number
        ro_number = data_list[0][11]
        hmsr = ""
        for data in cash_collection_data:
            if ro_number in data:
                hmsr = data[2]
        ## Customer information
        try:
            customer_name = data_list[4][1]
            customer_contact_1 = data_list[4][3]
            customer_contact_2 = data_list[4][5]
        except Exception as e:
            if customer_name == None:
                print("Null User Name")
            else:
                print("Check Customer Name")
            print(ro_number, e)
        finally:
            pass
        try:
            customer_type = data_list[4][7].lower()
        except Exception as e:
            print(e)
            print(ro_number)
        try:
            if customer_type == "private":
                customer_type = "Individual"
            elif customer_type == "Embassy":
                customer_type == "Individual"
            elif "IKLM" in customer_name:
                customer_type = "IKLM"
            elif "GKML" in customer_name:
                customer_type = "GKLM"
        except Exception as e:
            print(e,ro_number)
        finally:
            pass
        service_purpose = ""
        service_type = ""
        repair_type = ""
        service_reason = ""
        # os.system('cls')
        retail_part = 0
        retail_labour = 0
        part_discount = 0
        labour_discount = 0
        part_tax_percent = 0
        labour_tax_percent = 0
        subtotal = 0
        try:
            if data_list[15][10] != None:
                advisor_solve = data_list[15][10]
            elif data_list[14][10] != None:
                advisor_solve = data_list[14][10]
            else:
                advisor_solve = None
        except Exception as e:
            advisor_solve = None      
        # progress_bar.update(1)
        if "part" in excel_file.lower():
            try:
                advisor_solve = data_list[15][10]
                advisor_solve = os.linesep.join([s for s in advisor_solve.splitlines() if s])
            except Exception as e:
                print(e)
                print(data_list[15][10],"Advisor Solve")
                advisor_solve = ""   
            service_type = "Auto Parts"
            financial_data = calculate_final_total(service_type,excel_file)
            retail_part = financial_data[0]
            retail_labour = financial_data[1]
            part_discount = financial_data[2]
            labour_discount = financial_data[3]
            part_tax_percent = financial_data[4]
            labour_tax_percent = financial_data[5]
            subtotal = financial_data[6]
            sequence += 1
            print(sequence,".",ro_number,service_type,financial_data)
            ws.append([ro_number, car_takein_date, car_takein_time, car_takeout_date, car_takeout_time,"",1,
           "","","","","","","","",customer_type, customer_name, customer_contact_1,"Part", "Auto Parts","Buy Spare Parts From RO_2","","","",ro_number.replace("R","P"),ro_number.replace("R","V"),hmsr,
           retail_part, retail_labour, part_discount, labour_discount, part_tax_percent, labour_tax_percent,None, None,subtotal,last_value,0,last_value,"CR",None,"No",None,None,"Mr. Kyi Soe","A",advisor_solve])
        else:
            try:
                if data_list[15][10] != None:
                    advisor_solve = data_list[15][10]
                elif data_list[14][10] != None:
                    advisor_solve = data_list[14][10]
                else:
                    advisor_solve = None
            except Exception as e:
                advisor_solve = None 
            if data_list[8][7] == "Lastest Repair":
                car_name = data_list[7][2]
                car_model_year = data_list[7][5]
                try:
                    car_model_year = int(car_model_year)
                except Exception as e:
                    car_model_year = ""
                    pass
                car_color = data_list[7][6]
                car_plate_no = data_list[7][7]
                car_vin_no = data_list[7][8]
                car_mile_age = data_list[7][10]
                car_maker = ""
                car_type = ""
                try:
                    if car_model_year == "":
                        car_type = ""
                    else:
                        if car_model_year < 2014:
                            car_type = "Used"
                        else:
                            car_type = "New"
                except Exception as e:
                    print(e)

                    try:
                        if car_vin_no.startswith("K"):
                            car_maker = "CBU"
                        elif car_vin_no.startswith ("R"):
                            car_maker = "CBU"
                        else:
                            car_maker = "SKD"
                    except Exception as e:
                        continue
                        # print(excel_file, e) 

                    ## Reason
            else:
                car_name = data_list[8][2]
                car_model_year = data_list[8][5]
                car_color = data_list[8][6]
                car_plate_no = data_list[8][7]
                car_vin_no = data_list[8][8]
                car_mile_age = data_list[8][10]
                car_maker = ""
                car_type = ""
                try:
                    if car_model_year == "":
                        car_type = ""
                    else:
                        if int(car_model_year) < 2014:
                            car_type = "Used"
                        else:
                            car_type = "New"
                except Exception as e:
                    pass
                
                    # print(f"Please check this RO, {ro_number}. Car Model Year is somwthing wrong with the Error : {e}")
                finally:
                    pass
                try:
                    if car_vin_no.startswith("K"):
                        car_maker = "CBU"
                    elif car_vin_no.startswith ("R"):
                        car_maker = "CBU"
                    else:
                        car_maker = "SKD"
                except Exception as e:
                    continue

            service_reason = data_list[15][7]
            ## Assorted Service Types
            if advisor_solve == None:
                advisor_solve = "Blank"
            else:
                pass
            if "quotation" in advisor_solve.lower() or "qo" in excel_file.lower():
                service_type = "Quotation"
                ro_number = ro_number.replace("R","Q")

            elif "inspection" in advisor_solve.lower():
                service_type = "Inspection"
            elif "warranty" in advisor_solve.lower():
                service_purpose = "Repair"
                service_type = "Warranty"
                repair_type = "Break Down"
            elif "normal service" in advisor_solve.lower():
                service_purpose = "Repair"
                service_type = "General"
                repair_type = "Maintenance"
            else:
                service_purpose = "Repair"
                service_type = "General"
                repair_type = "Break Down"
            if service_type == "Quotation":
                sequence += 1
                print(sequence,".",ro_number, service_type)      
                ws.append([ro_number, car_takein_date, car_takein_time, car_takeout_date, car_takeout_time, pending,lead_day,
                   car_name, car_type, car_maker, car_plate_no, car_vin_no, car_mile_age, car_color, car_model_year,
                   customer_type, customer_name, customer_contact_1, service_purpose, service_type,
               service_reason, repair_type,"",ro_number,"","","",retail_part, retail_labour, part_discount, labour_discount, part_tax_percent, labour_tax_percent,None, None,None,0,0,0,"QO",None,None,None,None,"Mr. Kyi Soe","A",advisor_solve])
            elif pending == "Pending":
                sequence += 1
                print(sequence,".",ro_number,service_type)
                ws.append([ro_number, car_takein_date, car_takein_time, car_takeout_date, car_takeout_time, pending,lead_day,
                   car_name, car_type, car_maker, car_plate_no, car_vin_no, car_mile_age, car_color, car_model_year,
                   customer_type, customer_name, customer_contact_1,"", pending,
               service_reason,"","","","","","",retail_part, retail_labour, part_discount, labour_discount, part_tax_percent, labour_tax_percent,None, None,None,None,0,None,"Pending",None,None,None,None,"Mr. Kyi Soe","A",advisor_solve])

            elif service_type == "General":
                financial_data = calculate_final_total(service_type,excel_file)
                retail_part = financial_data[0]
                retail_labour = financial_data[1]
                part_discount = financial_data[2]
                labour_discount = financial_data[3]
                part_tax_percent = financial_data[4]
                labour_tax_percent = financial_data[5]
                subtotal = financial_data[6]
                sequence +=1
                print(sequence,".",ro_number,service_type,financial_data)
                ws.append([ro_number, car_takein_date, car_takein_time, car_takeout_date, car_takeout_time, pending,lead_day,
                   car_name, car_type, car_maker, car_plate_no, car_vin_no, car_mile_age, car_color, car_model_year,
                   customer_type, customer_name, customer_contact_1, service_purpose, service_type,
               service_reason, repair_type,"","",ro_number.replace("R","P"),ro_number.replace("R","V"),hmsr,retail_part, retail_labour, part_discount, labour_discount, part_tax_percent, labour_tax_percent,None, None,subtotal,last_value,0,last_value,"CR",None,"No",None,None,"Mr. Kyi Soe","A",advisor_solve])
            elif service_type == "Warranty":
                financial_data = calculate_final_total(service_type,excel_file)
                retail_part = financial_data[0]
                retail_labour = financial_data[1]
                part_discount = financial_data[2]
                labour_discount = financial_data[3]
                part_tax_percent = financial_data[4]
                labour_tax_percent = financial_data[5]
                subtotal = financial_data[6]
                sequence += 1
                print(sequence,".",ro_number,service_type, financial_data)
                ws.append([ro_number, car_takein_date, car_takein_time, car_takeout_date, car_takeout_time, pending,lead_day,
                   car_name, car_type, car_maker, car_plate_no, car_vin_no, car_mile_age, car_color, car_model_year,
                   customer_type, customer_name, customer_contact_1, service_purpose, service_type,
               service_reason, repair_type,"","","","","",retail_part, retail_labour, part_discount, labour_discount, part_tax_percent, labour_tax_percent,None, None,subtotal,last_value,0,last_value,"WTY",None,None,None,None,"Mr. Win Tun Oo","Additional Team",advisor_solve])
            
            elif service_type == "Inspection":
                sequence += 1
                print(sequence,".",ro_number,service_type)
                ws.append([ro_number, car_takein_date, car_takein_time, car_takeout_date, car_takeout_time, pending,lead_day,
                   car_name, car_type, car_maker, car_plate_no, car_vin_no, car_mile_age, car_color, car_model_year,
                   customer_type, customer_name, customer_contact_1, service_purpose, service_type,
               service_reason, repair_type,"","","","","",retail_part, retail_labour, part_discount, labour_discount, part_tax_percent, labour_tax_percent,None, None,subtotal,last_value,0,last_value,"FOC",None,None,None,None,"Mr. Kyi Soe","A",advisor_solve])
            
# progress_bar.close()          
# Save the workbook to a file
output_file_path = os.path.join(folder_path, 'ExtractedDataFromRO.xlsx')
new_file.save(output_file_path)
print("Output file saved successfully at:", output_file_path)


def list_excel_files(directory_path):
    try:
        # List all files in the directory
        files = os.listdir(directory_path)
        
        # Filter files to include only those with the '.xlsx' extension
        excel_files = [file.split('.')[0] for file in files if file.endswith('.xlsx')]
        
        # Return the list of Excel file names without the extension
        return excel_files
    except FileNotFoundError:
        print("Directory not found.")
        return []

def row_is_empty(row):
    return all(value is None for value in row)

# Load the Excel file into a pandas DataFrame, skipping the first column and the first row
excel_file_path = r"D:\Working Enviroment\Database\Data_2024_v2.2.xlsx"
sheet_name = "IKLM_data"
df = pd.read_excel(excel_file_path, sheet_name=sheet_name, skiprows=1)

# Remove the first column
df = df.iloc[:, 1:]

# Filter the DataFrame to only include rows where 'Pending Cars' column has value 'Pending'
pending_df = df[df['Pending Cars'] == 'Pending']

# Convert the 'RO No' column to a list
ro_number_list = pending_df["RO No"].tolist()

# Prompt the user to input the folder path
# directory_path = input("Enter the folder path: ")
directory_path = rf"Z:\IKLM_ASD\Reception\Mr.LEE\Pending"

# Validate the folder path
if not os.path.exists(directory_path):
    print("Invalid directory path.")
    exit()

# Get the list of Excel files with the '.xlsx' extension
excel_files = list_excel_files(directory_path)
sequence = 0

# Open a text file for writing the output in the same directory path
output_file_path = os.path.join(folder_path , 'Pending&FinishedList.txt')
with open(output_file_path, 'w') as f:
    for ro_number in ro_number_list:
        ro_found = False
        for file in excel_files:
            if ro_number in file:
                sequence += 1
                if 'finished' in file.lower():
                    output_text = f"{sequence} {ro_number} Finished\n"
                    print(output_text.strip())
                    f.write(output_text)
                else:
                    output_text = f"{sequence} {ro_number} Pending\n"
                    print(output_text.strip())
                    f.write(output_text)
                ro_found = True
                break
        if not ro_found:
            output_text = f"RO number {ro_number} not found in folder.\n"
            print(output_text.strip())
            f.write(output_text)

print("Pending Check List Output written to", output_file_path)

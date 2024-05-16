import os
from openpyxl import load_workbook
from openpyxl import Workbook

def extract_ro_and_reason(folder_path, output_file_path):
    file_list = os.listdir(folder_path)
    total_files = len([file for file in file_list if file.endswith(".xlsx")])
    
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.append(["RO or QO Number", "Reason"])

    processed_files = 0

    for file_name in file_list:
        excel_file = os.path.join(folder_path, file_name)
        if not excel_file.endswith(".xlsx"):
            continue
        else:
            wb = load_workbook(excel_file)
            sheet_name = "Repair_Order_01"
            sheet = wb[sheet_name]

            # Find the index of the row containing "RO No.:"
            ro_row_index = None
            for i, row in enumerate(sheet.iter_rows()):
                for cell in row:
                    if "RO No.:" in str(cell.value):
                        ro_row_index = i
                        break
                if ro_row_index is not None:
                    break

            if ro_row_index is not None:
                # Delete rows above the row containing "RO No.:"
                sheet.delete_rows(1, ro_row_index - 1)

            # Define Column Range
            column_range = "A:M"
            data_list = []

            # Iterate through rows and extract data from columns A to M
            for row in sheet.iter_rows(min_col=1, max_col=13, values_only=True):
                if any(row):
                    data_list.append(row)

            ro_number = data_list[0][11]  # Extract RO or QO number
            service_reason = data_list[15][7]  # Extract Reason

            # Append the extracted data to the new sheet
            new_sheet.append([ro_number, service_reason])

            processed_files += 1
            progress_percentage = (processed_files / total_files) * 100
            print(f"Progress: {progress_percentage:.2f}% ({processed_files}/{total_files})", end="\r")

    # Save the new workbook to the specified output file path
    try:
        new_workbook.save(output_file_path)
        print("\nData recorded successfully. Output file saved at:", output_file_path)
    except Exception as e:
        print("\nAn error occurred while saving the output file:", e)

# Prompt the user to input the folder path
folder_path = input("Enter the folder path: ")

# Validate the folder path
if not os.path.exists(folder_path):
    print("Invalid folder path.")
else:
    # Prompt the user to input the output file path
    output_file_path = input("Enter the output file path (e.g., output.xlsx): ")
    extract_ro_and_reason(folder_path, output_file_path)

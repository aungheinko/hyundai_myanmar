from openpyxl import load_workbook, Workbook

# Load the input Excel file
input_file = rf"C:\Users\asservices012\Desktop\hello.xlsx"  # Replace with your input file name
output_file = "output.xlsx"  # Replace with your desired output file name

wb = load_workbook(input_file)
ws = wb.active

# Create a new workbook for the output
output_wb = Workbook()
output_ws = output_wb.active
output_ws.append(["Part No", "Description", "Take-In Date", "Take-In Qty", "Comment"])  # Header

# Iterate through the rows in the input file
for row in ws.iter_rows(min_row=2):  # Start from the second row to skip headers
    part_no = row[0].value  # Column A: Part No
    description = row[1].value  # Column B: Description

    for col_idx in range(2, len(row)):  # Start from column C (dates)
        cell = row[col_idx]
        date = ws.cell(row=1, column=col_idx + 1).value  # Get the date from the header row

        if cell.value is not None:  # Check if there's a value in the cell
            comment = cell.comment.text if cell.comment else ""  # Get the comment if available
            # Append to the output file
            output_ws.append([part_no, description, date, cell.value, comment])

# Save the output file
output_wb.save(output_file)
print(f"Data successfully written to {output_file}")
